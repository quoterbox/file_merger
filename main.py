import os
from openpyxl import load_workbook
import csv


def get_data_from_xlsx(xlsx_name: str) -> []:
    wb = load_workbook(filename=xlsx_name)
    ws = wb.active
    xlsx_rows = []
    cnt = 0

    for row in ws.values:
        one_row = create_dict_fields(row)
        if cnt > 0:
            xlsx_rows.append(one_row)
        cnt += 1

    return xlsx_rows


def get_data_from_csv(csv_name: str, fieldnames: {}) -> []:
    csv_rows = []
    with open(csv_name, "r") as doc:
        reader = csv.DictReader(doc, delimiter=";", fieldnames=fieldnames)
        next(reader, None)
        for row in reader:
            csv_rows.append(row)

    return csv_rows


def get_data_from_file(file_name: str, fieldnames: {}):
    if file_name.endswith(".xlsx"):
        return get_data_from_xlsx(file_name)
    elif file_name.endswith(".csv"):
        return get_data_from_csv(file_name, fieldnames)


def get_csv_first_line(csv_name: str) -> []:
    with open(csv_name, "r", newline="") as doc:
        reader = csv.DictReader(doc, delimiter=";")
        headers = reader.fieldnames

    return headers


def get_xlsx_first_line(xlsx_name: str) -> []:
    wb = load_workbook(filename=xlsx_name)
    ws = wb.active
    row = ""

    for row in ws.values:
        break

    return row


def get_file_first_line(file_name: str) -> []:
    if file_name.endswith(".xlsx"):
        return get_xlsx_first_line(file_name)
    elif file_name.endswith(".csv"):
        return get_csv_first_line(file_name)


def create_target_file(file_name: str, line: {}):
    with open(file_name, "a", newline="") as doc:
        writer = csv.DictWriter(doc, delimiter=";", fieldnames=line)
        writer.writerow(line)


def create_dict_fields(row: {}) -> {}:
    try:
        keys = range(len(row))
        return dict(zip(keys, row))
    except 'NoneType':
        print("Row has NoneType - %s" % row)


# directory_name = "files_csv"
directory_name = "files_xlsx"
target_file = "merged_files.csv"
fields = {}
cnt = 0

for file in os.listdir(directory_name):
    relpath = os.path.join(directory_name, file)

    if relpath.endswith(".xlsx") or relpath.endswith(".csv"):
        if cnt == 0:
            first_line = get_file_first_line(relpath)
            fields = create_dict_fields(first_line)
            create_target_file(target_file, fields)

        rows = get_data_from_file(relpath, fields)

        with open(target_file, "a", newline="") as tg_file:
            writer = csv.DictWriter(tg_file, delimiter=";", fieldnames=fields)
            for row in rows:
                writer.writerow(row)
        cnt += 1
