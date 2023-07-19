import os
from openpyxl import load_workbook
import csv


def create_dict_fields(row: {}) -> {}:
    try:
        keys = range(len(row))
        return dict(zip(keys, row))
    except 'NoneType':
        print("Row has NoneType - %s" % row)


class TableFile:
    _file_name = ""

    def __init__(self, file_name: ""):
        self._file_name = file_name

    def get_file_first_line(self) -> []:
        return []

    def get_data_from_file(self) -> []:
        return []


class XslxFile(TableFile):

    def get_file_first_line(self) -> []:
        wb = load_workbook(filename=self._file_name)
        ws = wb.active
        row = ""

        for row in ws.values:
            break

        return row

    def get_data_from_file(self) -> []:
        wb = load_workbook(filename=self._file_name)
        ws = wb.active
        xlsx_rows = []
        cnt = 0

        for row in ws.values:
            if cnt > 0:
                xlsx_rows.append(row)
            cnt += 1

        return xlsx_rows


class CsvFile(TableFile):
    __encoding = ""
    __delimiter = ";"
    __newline = ""

    def __init__(self, file_name: str, delimiter: str, newline: str, **kwargs):
        super().__init__(file_name)
        self.__delimiter = delimiter
        self.__encoding = kwargs.get("encoding", "utf-8")
        self.__newline = newline

    def get_file_first_line(self) -> []:
        with open(self._file_name, "r", encoding=self.__encoding, newline=self.__newline) as doc:
            reader = csv.DictReader(doc, delimiter=self.__delimiter)
            headers = reader.fieldnames
        return headers

    def get_data_from_file(self) -> []:
        csv_rows = []
        with open(self._file_name, "r", encoding=self.__encoding, newline=self.__newline) as doc:
            reader = csv.reader(doc, delimiter=self.__delimiter)
            next(reader, None)
            for row in reader:
                csv_rows.append(row)
        return csv_rows


class FMerger:
    valid_files = [".csv", ".xlsx"]
    table_file = None
    source_dir = ""
    source_encoding = "utf-8"
    source_delimiter = ";"
    source_newline = ""
    target_file = "merged_files.csv"
    target_encoding = "utf-8"
    target_delimiter = ";"
    target_newline = ""

    def __init__(self, options: {}):
        self.source_dir = options["sources_options"]["source_dir"]
        if "source_encoding" in options.get("sources_options", {}):
            self.source_encoding = options["sources_options"]["source_encoding"]
        self.source_delimiter = options["sources_options"]["source_delimiter"]
        self.source_newline = options["sources_options"]["source_newline"]
        self.target_file = options["target_file_options"]["target_file"]
        if "target_encoding" in options.get("target_file_options", {}):
            self.target_encoding = options["target_file_options"]["target_encoding"]
        self.target_delimiter = options["target_file_options"]["target_delimiter"]
        self.target_newline = options["target_file_options"]["target_newline"]

    def merge(self):
        cnt = 0
        for file in self.__read_directory_files(self.source_dir):
            table_file = self.__create_table_file(file, self.source_delimiter, self.source_newline, self.source_encoding)
            if cnt == 0:
                first_line = table_file.get_file_first_line()
                self.__create_target_file(self.target_file, create_dict_fields(first_line))

            rows = table_file.get_data_from_file()
            self.__write_target_file(rows)
            cnt += 1

    def __create_target_file(self, file_name: str, line: {}):
        with open(file_name, "a", encoding=self.target_encoding, newline=self.target_newline) as doc:
            writer = csv.DictWriter(doc, delimiter=self.target_delimiter, fieldnames=line)
            writer.writerow(line)

    def __write_target_file(self, rows: []):
        with open(self.target_file, "a", encoding=self.target_encoding, newline=self.target_newline) as tg_file:
            writer = csv.writer(tg_file, delimiter=self.target_delimiter)
            for row in rows:
                writer.writerow(row)

    def __is_valid_file(self, file_extension: str) -> bool:
        return file_extension in self.valid_files

    def __read_directory_files(self, directory_name: str) -> []:
        files = []

        for file in os.listdir(directory_name):
            file_path = os.path.join(directory_name, file)
            filename, file_extension = os.path.splitext(file)
            if self.__is_valid_file(file_extension):
                files.append(file_path)

        return files

    @staticmethod
    def __create_table_file(table_file_path: str, delimiter: str, newline: str, encoding: str) -> TableFile:
        table = None
        if table_file_path.endswith(".xlsx"):
            table = XslxFile(table_file_path)
        elif table_file_path.endswith(".csv") and delimiter:
            table = CsvFile(table_file_path, delimiter, newline, encoding=encoding)
        return table
